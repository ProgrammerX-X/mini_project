from openpyxl import Workbook, load_workbook
import sqlite3 as sq
import os

db_name = "projects.db"
way = os.path.join("..\\mini_project\\databases\\" + db_name)
db = sq.connect(way)
cursor = db.cursor()

excel_file = "projects.xlsx"
way_for_excel = os.path.join("..\\mini_project\\databases\\" + excel_file)

try:
    workbook = load_workbook(way_for_excel)
except FileNotFoundError:
    workbook = Workbook()

def generate_standarts(excel_file):
    try:    
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        workbook = Workbook()
        
        sheet_data = {
            "PROJECTS": ["NUMBER_OF_PROJECT", "NAME", "NUM_OF_CUSTOMER",
                        "PROJECT_MANAGER_NUM", "DURATION_OF_EXECUTION", "DIFFICULTY_CATEGORY", "BEGIN_DATE"],
            "ABOUT_COMPANY": ["PERFORMER_NUMBER", "F_L_P", "POST", "SKILL_LEVEL"],
            "PROJECTS_IN_PROGRESS": ["NUMBER_OF_PROJECT", "PERFORMER_NUMBER"],
            "COEFFICIENTS": ["DIFFICULTY_NUMBER", "DIFFICULTY_PROJECT", "ALLOWANCE_COEFFICIENT"],
            "PERFORMERS": ["PERFORMER_NUMBER", "CUSTOMER"],
            "PAYROLL": ["PERFORMER_NUMBER", "BANK_ACCOUNT_NUMBER", "WAGE"],
            "DEBTORS": ["NUMBER_OF_DEBTORS", "CUSTOMER_NUMBER", "AMOUNT_OF_DEBT"],
            "PAYMENT": ["POST", "HOURLY_RATE_PAYMENT", "ALLOWANCE_RATIO"],
            "REPORT": ["REPORT_NUMBER", "PROJECT", "CHANGES", "HOURS"]
        }

        for sheet_name, headers in sheet_data.items():
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(headers)

        workbook.save(excel_file)

# generate_standarts(way_for_excel, workbook)

def fake_datas(excel_file):
    workbook = load_workbook(excel_file)
    sheet_name = "PROJECTS"

    sheet = workbook[sheet_name]

    experemental_datas = [
        [1, "Construction Calculator", 1001, 3016, 3, 2, "01.02.2025"],
        [2, "Reservation system", 1002, 3017, 6, 2, "01.02.2025"],
        [3, "BuildPro", 1003, 3018, 5, 2, "16.05.2025"],
        [4, "Online store with elements gamification", 1004, 3019, 8, 3, "30.11.2025"],
        [5, "Marketplace for rent expensive possession", 1005, 3020, 7, 3, "31.10.2025"],
        [6, "Shared Office 2.0 platform", 1006, 3021, 6, 2, "16.07.2026"],
        [7, "Virtual showroom for business", 1007, 3022, 5, 2, "16.06.2026"],
        [8, "Exchange for job exchanges", 1008, 3023, 4, 2, "28.02.2027"],
        [9, "Hobby monetization service", 1009, 3024, 6, 2, "31.01.2027"],
        [10, "Micro-courses for experts", 1010, 3025, 7, 3, "16.06.2027"],
        [11, "Eco-friendly marketplace for business", 1011, 3026, 3, 2, "16.07.2027"],
        [12, "VR for viewing real estate", 1012, 3027, 12, 4, "28.02.2028"],
        [13, "Platform for testing ideas", 1013, 3028, 5, 2, "31.01.2028"],
        [14, "Startup in a box", 1014, 3029, 3, 2, "28.02.2029"]
    ]

    existing_projects = {row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)}

    for row in experemental_datas:
        if row[0] not in existing_projects:
            sheet.append(row)

    workbook.save(excel_file)

# fake_datas(way_for_excel)

def getter(excel_file):
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        generate_standarts(excel_file)
        workbook = load_workbook(excel_file)

    data_work = ["PROJECTS", "ABOUT_COMPANY", "PROJECTS_IN_PROGRESS", 
                 "COEFFICIENTS", "PERFORMERS", "PAYROLL", 
                 "DEBTORS", "PAYMENT", "REPORT"]
    comp = []
    for k in data_work:
        try:
            sheet = workbook[k]
        except KeyError:
            generate_standarts(excel_file)
            sheet = workbook[k]
        dt = []
        for row in sheet.iter_rows(values_only=True):
            dt.append(row)
        # comp.append(dt)
    print(dt)
    # return comp

getter("p.xlsx")
# print(p)

def comparison(old_file, file):
    old_f = getter(old_file)
    new_f = getter(file)

    workbook = load_workbook(file)

    data_work = ["PROJECTS", "ABOUT_COMPANY", "PROJECTS_IN_PROGRESS", 
                 "COEFFICIENTS", "PERFORMERS", "PAYROLL", 
                 "DEBTORS", "PAYMENT", "REPORT"]
    
    # for i in old_f:
    #     print(i)
    # print("\n\n\n\n\n")

    # for j in new_f:
    #     print(j)



    # for sheet_name in data_work:
        # if sheet_name in workbook.sheetnames:
        #     sheet = workbook[sheet_name]
            
        # else:
        #     generate_standarts(file)
        #     workbook = load_workbook(file)
        #     sheet = workbook[sheet_name]

        # old_sheet_data = old_f[data_work.index(sheet_name)]
        # new_sheet_data = new_f[data_work.index(sheet_name)]
        # print(sheet_name)
    # for i in old_sheet_data:
    #     print(i)
    # print("\n\n\n\n\n\n\n")
    # for i in new_sheet_data:
    #     print(i)

    for i, rows_idx in zip(data_work, range(len(data_work))):
        sheet = workbook[i]
        for old_v, cols_idx in zip(old_f, range(len(old_f))):
            # if new_v != old_v:
                print(old_v)
            #     # sheet.cell(row=rows_idx+1,column=cols_idx+1,value=old_)
            # else:
            #     continue

    workbook.save(file)

# comparison(way_for_excel, "p.xlsx")